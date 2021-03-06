import openpyxl
import configparser
import csv
import os
from numpy import median
from numpy import average
from numpy import transpose
from copy import copy

# Config vars

config = configparser.ConfigParser()
config.read('scriptconfig.ini')

# Required input vars
PROTOCOL_WB = config['run info']['protocol_wb']
OUTPUT_WB = config['run info']['output_wb']

def check_and_default(config,cat,key,default):
    if key in config[cat].keys():
        return config[cat][key]
    else:
        return default

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

# Run info
NUM_INPUT = int(check_and_default(config,'run info','num_input','20')) # number of samples
NUM_BLOCKS = int(check_and_default(config,'run info','num_blocks','16')) # number of blocks on slide
SAVE_ENABLED = check_and_default(config,'run info','save_enabled','True') == 'True'
BLOCKWISE_PBS = check_and_default(config,'run info','blockwise_pbs','True') == 'True'
MULTIBLOCK_ANTIGENS = check_and_default(config,'run info','multiblock_antigens','False') == 'True'
BLANK_SAMPLES = check_and_default(config,'run info','blank_samples','True') == 'True'

# Results File Info
DATA_COL = check_and_default(config,'results file','data_col','Z')
FLAG_COL = check_and_default(config,'results file','flag_col','A')
NAME_COL = check_and_default(config,'results file','name_col','G')
BLOC_COL = check_and_default(config,'results file','bloc_col','D')
FIRST_ROW_DATA = int(check_and_default(config,'results file','first_row_data','34'))
AUTO_COL = check_and_default(config,'results file','auto_col','False') == 'True'

# Protocol File Info
SAMPLE_COL = check_and_default(config,'protocol file','sample_col','B')
SECOND_COL = check_and_default(config,'protocol file','second_col','E')
SAMPLE_ROW = int(check_and_default(config,'protocol file','sample_row','20'))

# Debug vars
VERBOSE_OUTPUT = check_and_default(config,'debug','verbose_output','False') == 'True'

# ===== FUNCTIONS ======

# returns True if variable can be cast as Float
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
    except TypeError:
        return False

# creates copy of sheet [num] in specified workbook
# sets all digit values < 1 to 1
# inserts new sheet after copy of specified sheet with name + " floored"
def sheetfloor(wb,num):
    og = wb.worksheets[num]
    ws = wb.create_sheet(og.title+" floored", num+1)
    for j in range(len(og['A'])):
        for i in range(len(og[1])):
            value = og.cell(row = j+1, column = i+1).value
            font = copy(og.cell(row = j+1, column = i+1).font)
            ws.cell(row = j+1, column = i+1).font = font
            if is_number(value) and float(value) < 1:
                ws.cell(row = j+1, column = i+1).value = 1
            else:
                ws.cell(row = j+1, column = i+1).value = value

#returns letter name of column of given count
def numtocol(num):
    out = ""
    while(num>0):
        digit = (num - 1) % 26
        out = chr(digit+65) + out
        num = int((num-1)/26)
    return out

# ===== SCRIPT =====

# Set up WB's
wb_protocol = openpyxl.load_workbook(PROTOCOL_WB)
os.chdir("Results File")

# Combine all data into one .xlsx book with each sample on a page

wb_combined = openpyxl.Workbook()
for i in range(NUM_INPUT):
    if(i > len(wb_combined.sheetnames) - 1):
       wb_combined.create_sheet()
    print("loading sheet " + str(i))
    ws = wb_combined.worksheets[i]
    curnum = str(i+1).zfill(len(str(NUM_INPUT)))
    # everything else is 1-indexed and 0-padded

    ws.title = curnum
    with open("slide"+curnum+".txt") as data:
        reader = csv.reader(data, delimiter='\t')
        for row in reader:
            ws.append(row)
if(SAVE_ENABLED):
    print("saving data_combined.xlsx...")
    wb_combined.save("data_combined.xlsx")
    print("data_combined.xlsx saved")

os.chdir("..")

# create one sheet in a book with all relevant data, flagged samples removed

wb_working = openpyxl.Workbook()
ws = wb_working.worksheets[0]
ws.title = "raw medians"


for i in range(len(wb_combined.worksheets)):
    currsheet = wb_combined.worksheets[i]
    if(VERBOSE_OUTPUT):
        print("condensing sheet " + str(i))
    j = 1 #side note: I hate 1-indexing
    # get ready to see it a lot

    #making local var versions of input columns
    data_col = DATA_COL
    name_col = NAME_COL
    flag_col = FLAG_COL
    bloc_col = BLOC_COL

    #override local col vars if told to look for it, needs to be unique per sheet
    if(AUTO_COL):
        if(VERBOSE_OUTPUT):
            print("automatically finding proper column headers")
        col_test = 1
        cols_set = 0
        while(cols_set < 4):
            header = currsheet[numtocol(col_test)+str(FIRST_ROW_DATA)].value
            if(header == "F532 Median - B532"):
                data_col = numtocol(col_test)
                if(VERBOSE_OUTPUT):
                    print("data_col found: " + data_col)
                cols_set = cols_set + 1
            elif(header == "Name"):
                name_col = numtocol(col_test)
                if(VERBOSE_OUTPUT):
                    print("name_col found: " + name_col)
                cols_set = cols_set + 1
            elif(header == "Flags"):
                flag_col = numtocol(col_test)
                if(VERBOSE_OUTPUT):
                    print("flag_col found: " + flag_col)
                cols_set = cols_set + 1
            elif(header == "Block"):
                bloc_col = numtocol(col_test)
                if(VERBOSE_OUTPUT):
                    print("bloc_col found: " + bloc_col)
                cols_set = cols_set + 1
            col_test = col_test + 1
        
    while(currsheet[data_col+str(j+FIRST_ROW_DATA)].value is not None):
        if(i==0):
            clean_name = currsheet[name_col+str(j+FIRST_ROW_DATA)].value.replace("_","-")
            if VERBOSE_OUTPUT and clean_name != currsheet[name_col+str(j+FIRST_ROW_DATA)].value:
                print("Changed '"+ currsheet[name_col+str(j+FIRST_ROW_DATA)].value +"' to '"+ clean_name +"'")
            ws['A'+str(j+1)] = clean_name + "_" + (currsheet[bloc_col+str(j+FIRST_ROW_DATA)].value)
        if(currsheet[flag_col+str(j+FIRST_ROW_DATA)].value == '-100'):
            ws.cell(row = j+1, column = i+2).value = 'NA'
        else:
            ws.cell(row = j+1, column = i+2).value = currsheet[data_col+str(j+FIRST_ROW_DATA)].value
        j = j+1

# add sample names for column titles

if(VERBOSE_OUTPUT):
    print("Adding sample names from Protocol File")
for i in range(NUM_INPUT):
    ws.cell(row = 1, column = i+2).value = wb_protocol['Protocol'][SAMPLE_COL + str(SAMPLE_ROW+i)].value

# performing median confinement

print("Consolidating identical analytes")
data = dict()
for i in range(len(ws['A'])-1):
    key = ws['A'+str(i+2)].value
    if key not in data.keys():
        data[key] = [list() for a in range(NUM_INPUT)]
        if(VERBOSE_OUTPUT):
            print("Adding analyte ID "+key+"...")
    for j in range(NUM_INPUT):
        data[key][j].append(ws.cell(row = i+2, column = j+2).value)

if(VERBOSE_OUTPUT):
    print("Creating new worksheet")
ws = wb_working.create_sheet("median medians")

if(VERBOSE_OUTPUT):
    print("Adding sample and secondary names from Protocol File")
for i in range(NUM_INPUT):
    clean_name = wb_protocol['Protocol'][SAMPLE_COL + str(SAMPLE_ROW+i)].value.replace("_","-")
    if VERBOSE_OUTPUT and clean_name != wb_protocol['Protocol'][SAMPLE_COL + str(SAMPLE_ROW+i)].value:
        print("Changed '"+ wb_protocol['Protocol'][SAMPLE_COL + str(SAMPLE_ROW+i)].value +"' to '"+ clean_name +"'")
    ws.cell(row = 1, column = i+2).value = clean_name + "_" + wb_protocol['Protocol'][SECOND_COL + str(SAMPLE_ROW+i)].value

print("Calculating median values")
i = 2
for key in data.keys():
    ws.cell(row = i, column = 1).value = key
    if(VERBOSE_OUTPUT):
        print("Finding median of "+key+"...")
    for j in range(NUM_INPUT):
        values = [int(a) for a in data[key][j] if a != 'NA']
        if values == list():
            ws.cell(row = i, column = j+2).value = 'NA'
        else:
            ws.cell(row = i, column = j+2).value = median(values)
    i = i+1

if(VERBOSE_OUTPUT):
    print("Setting floor to 1")
sheetfloor(wb_working, 1)

print("Consolidating identical samples")
ws = wb_working.worksheets[2]

data = dict()
num_analytes = len(ws['A'])-1
for i in range(NUM_INPUT):
    key = ws.cell(row = 1, column = i+2).value
    if key not in data.keys():
        data[key] = [list() for a in range(num_analytes)]
        if(VERBOSE_OUTPUT):
            print("Adding sample ID "+key+"...")
    for j in range(num_analytes):
        data[key][j].append(ws.cell(row = j+2, column = i+2).value)

if(VERBOSE_OUTPUT):
    print("Creating new worksheet")
ws = wb_working.create_sheet("mean samples")

if(VERBOSE_OUTPUT):
    print("Adding analyte names from previous sheet")
for i in range(num_analytes):
    ws.cell(column = 1, row = i+2).value = wb_working.worksheets[2].cell(column = 1, row = i+2).value

print("Calculating average values")
i = 2
for key in data.keys():
    ws.cell(row = 1, column = i).value = key
    if(VERBOSE_OUTPUT):
        print("Averaging values for "+key+"...")
    for j in range(num_analytes):
        values = [float(a) for a in data[key][j] if a != 'NA']
        if values == list():
            ws.cell(row = j+2, column = i).value = 'NA'
        else:
            ws.cell(row = j+2, column = i).value = average(values)
    i = i+1

if(VERBOSE_OUTPUT):
    print("Separating block identifiers")
num_samples = len(ws[1]) - 1
data = [dict() for a in range(NUM_BLOCKS)]
for i in range(len(ws['A'])-1):
    ID = ws.cell(column = 1, row = i+2).value
    block = int(ID.split("_")[1]) - 1
    key = ID.split("_")[0]
    data[block][key] = [0 for a in range(num_samples)]
    if(VERBOSE_OUTPUT):
        print("Adding analyte ID "+key+" to block "+str(block)+"...")
    for j in range(num_samples):
        data[block][key][j] = ws.cell(row = i+2, column = j+2).value
        
if(BLOCKWISE_PBS):
    if(VERBOSE_OUTPUT):
        print("Creating new worksheet")
    ws = wb_working.create_sheet("PBS corrected")

    if(VERBOSE_OUTPUT):
        print("Adding sample names from previous sheet")
    for i in range(num_samples):
        ws.cell(column = i+2, row = 1).value = wb_working.worksheets[3].cell(column = i+2, row = 1).value

    currrow = 2
    for block in range(NUM_BLOCKS):
        for key in data[block].keys():
            ws.cell(column = 1, row = currrow).value = key
            for i in range(len(data[block][key])):
                curval = data[block][key][i]
                curPBS = data[block]["PBS"][i]
                if(curval=='NA'):
                    ws.cell(column = i+2, row = currrow).value = 'NA'
                elif(curPBS=='NA'):
                    # if the PBS value is invalid, leave as is and change format
                    ws.cell(column = i+2, row = currrow).value = curval
                    ws.cell(column = i+2, row = currrow).font = openpyxl.styles.Font(italic=True)
                else:
                    ws.cell(column = i+2, row = currrow).value = curval - curPBS
                    data[block][key][i] = curval - curPBS
            currrow = currrow + 1

    if(VERBOSE_OUTPUT):
        print("Setting floor to 1")
    sheetfloor(wb_working, len(wb_working.worksheets)-1)

if(MULTIBLOCK_ANTIGENS):
    if(VERBOSE_OUTPUT):
        print("Creating new worksheet")
        print("Averaging antigens across blocks")
        
    ws = wb_working.create_sheet("Blockwise Antigens Avg")
    
    per_sample = dict()
    for block in range(NUM_BLOCKS):
        for key in data[block].keys():
            if key in per_sample.keys():
                per_sample[key].append(data[block][key])
            else:
                per_sample[key] = [data[block][key]]


    if(VERBOSE_OUTPUT):
        print("Adding sample names from previous sheet")
    for i in range(num_samples):
        ws.cell(column = i+2, row = 1).value = wb_working.worksheets[3].cell(column = i+2, row = 1).value

    keynum = 0
    for key in per_sample.keys():
        ws.cell(row = keynum+2, column = 1).value = key
        
        rearr = transpose(per_sample[key])
        for i in range(len(rearr)):
            ws.cell(column = i+2, row=keynum+2).value = average([x.astype(float) for x in rearr[i] if is_number(x)])

        keynum = keynum + 1
    

if(BLANK_SAMPLES):
    if(VERBOSE_OUTPUT):
        print("Subtracting blanks from matching secondary")
    ws = wb_working.worksheets[-1]

    data = dict()
    styles = dict()
    for i in range(num_samples):
        ID = ws.cell(row = 1, column = i+2).value
        sample = ID.split("_")[0]
        secondary = ID.split("_")[1]
        if secondary not in data.keys():
            data[secondary]=dict()
            styles[secondary]=dict()
        for col in ws.iter_cols(min_row=2, min_col=i+2,max_col=i+2):
            data[secondary][sample]=[cell.value for cell in col]
            styles[secondary][sample]=[cell.font for cell in col] #need to be sure to copy styles at this point

    if(VERBOSE_OUTPUT):
        print("Creating new worksheet")
    ws = wb_working.create_sheet("blank subtracted")

    if(VERBOSE_OUTPUT):
        print("Adding analyte names from previous sheet")
    for i in range(num_analytes):
        ws.cell(column = 1, row = i+2).value = wb_working.worksheets[-2].cell(column = 1, row = i+2).value

    currcol = 2
    for secondary in data.keys():
        for sample in data[secondary].keys():
            ws.cell(row = 1, column = currcol).value = sample
            for i in range(len(data[secondary][sample])):
                curval = data[secondary][sample][i]
                curblank = data[secondary]['Blank'][i]
                if(curval=='NA'):
                    ws.cell(column = currcol, row = i+2).value = 'NA'
                elif(curblank=='NA'):
                    ws.cell(column = currcol, row = i+2).value = curval
                   # ws.cell(column = currcol, row = i+2).font = openpyxl.Font(bold=True)
                else:
                    ws.cell(column = currcol, row = i+2).value = curval - curblank
                ws.cell(column = currcol, row = i+2).font = copy(styles[secondary][sample][i])
            currcol = currcol + 1

    if(VERBOSE_OUTPUT):
        print("Setting floor to 1")
    sheetfloor(wb_working, len(wb_working.worksheets)-1)
        
if(SAVE_ENABLED):
    wb_working.save("Results_Normalization_Process.xlsx")  
    print("Saved file 'Results_Normalization_Process.xlsx'")

    while(len(wb_working.worksheets) > 1):
        wb_working.remove_sheet(wb_working.worksheets[0])
    wb_working.worksheets[0].title = "master"
    wb_working.save(OUTPUT_WB)  
    print("Saved file '" + OUTPUT_WB + "'")


